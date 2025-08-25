import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import webbrowser
import requests
import datetime
import csv
import os
import json
import math
import sys
import re
import subprocess
from bs4 import BeautifulSoup

# アプリバージョン
APP_VERSION = "2.07"

# アップデート確認
UPDATE_CHECK_URL = "https://sites.google.com/view/yo-payroll/latest"
UPDATE_PAGE_URL = "https://sites.google.com/view/yo-payroll/#h.6oue5jnv93fw"

# 除外する列名
EXCLUDED_COLUMNS = ['月', '日', '曜日', '空き', '予定', '.']

# アップデート確認関数
def check_for_updates():
    try:
        # 最新バージョンを確認
        response = requests.get(UPDATE_CHECK_URL, timeout=10)
        response.raise_for_status()

        # ページタイトルを取得
        soup = BeautifulSoup(response.text, 'html.parser')
        title = soup.title.string.strip()
        latest_version = title.split()[-1]

        # バージョン比較
        if float(latest_version) > float(APP_VERSION):
            # ページ内容を取得
            content = soup.get_text(separator="\n").strip()
            
            # <start>と<end>の間のテキストを抽出
            start_tag = "<start>"
            end_tag = "<end>"
            start_index = content.find(start_tag)
            end_index = content.find(end_tag, start_index)
            
            if start_index != -1 and end_index != -1:
                update_details = content[start_index + len(start_tag):end_index].strip()
            else:
                update_details = "更新内容の取得できませんでした。\n\nヘルプページで直接確認してください。"

            # 更新メッセージを作成
            update_message = (
                f"バージョン {latest_version} のアップデートがあります。\n\n"
                f"更新内容は以下の通りです。\n\n{update_details}\n\n"
                f"アップデートしますか？"
            )

            # 更新するかどうかを尋ねる
            if messagebox.askyesno("アップデート確認", update_message):
                subprocess.Popen(["C:\\payroll_app\\updater.exe"])  # アップデーターを起動
                root.destroy()  # Tkinterウィンドウを閉じる
                sys.exit()  # プログラムを終了

    except requests.RequestException as e:
        messagebox.showwarning("アップデート確認エラー", f"アップデート確認に失敗しました。\n{e}")
    except Exception as e:
        messagebox.showerror("アップデート確認エラー", f"予期せぬエラーが発生しました。\n{e}")

# ツールチップクラスの定義
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        if self.tip_window:
            return
        x, y, _, _ = self.widget.bbox("insert")  # ウィジェットの座標取得
        x += self.widget.winfo_rootx() + 25  # ウィンドウ位置調整
        y += self.widget.winfo_rooty() + 25
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)  # 枠なしウィンドウ
        self.tip_window.geometry(f"+{x}+{y}")
        label = tk.Label(self.tip_window, text=self.text, bg="lightyellow", relief="solid", borderwidth=1, padx=5, pady=2)
        label.pack()

    def hide_tooltip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


# GUI作成
root = tk.Tk()
root.title(f"簡易給料計算ソフト ver.{APP_VERSION}")
root.overrideredirect(False)

# 画面のサイズを取得
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# ウィンドウの幅と高さを取得（現在のサイズ）
root.update_idletasks()  # これでウィンドウサイズが正しく取得できる
win_width = root.winfo_width()
win_height = root.winfo_height()

# ウィンドウを中央に配置するためのX, Y座標を計算
x = (screen_width - win_width) // 2
y = (screen_height - win_height) // 3

# ウィンドウの位置を設定
root.geometry(f"+{x}+{y}")

# アップデート確認実行
root.after(100, check_for_updates)

# 設定ファイルのパス
SETTINGS_FILE = r"C:\payroll_app\settings.json"

# 設定を保存
def save_settings(base_wage, overtime_multiplier, adjustment_minutes):
    os.makedirs(os.path.dirname(SETTINGS_FILE), exist_ok=True)
    settings = {
        "base_wage": int(float(base_wage)),  # 整数として保存
        "overtime_multiplier": float(overtime_multiplier),
        "adjustment_minutes": int(adjustment_minutes)  # 追加
    }
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(settings, f)

# 設定を読み込み
def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                settings = json.load(f)
            if "base_wage" in settings and "overtime_multiplier" in settings:
                # もし adjustment_minutes が設定ファイルにない場合はデフォルト値を設定
                if "adjustment_minutes" not in settings:
                    settings["adjustment_minutes"] = 20
                return settings
            else:
                raise ValueError("設定ファイルが不正です")
        except (json.JSONDecodeError, ValueError):
            messagebox.showwarning("読み込みエラー", "保存された基本時給と時間外勤務倍率が読み込めませんでした。\nデフォルト値を使用します。")
            os.remove(SETTINGS_FILE)
    return {
        "base_wage": 1162,
        "overtime_multiplier": 1.25,
        "adjustment_minutes": 20  # デフォルト値を設定
    }


# Excelファイル読み込み関数
def load_sheet(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = workbook.sheetnames
    
    # シートが複数ある場合、ラジオボタンで選択ダイアログを表示
    if len(sheet_names) > 1:
        selection_window = tk.Toplevel(root)
        selection_window.title("シート選択")
        selection_window.overrideredirect(False)
        selection_window.geometry(f"+{x}+{y}")
        tk.Label(selection_window, text="データ処理するシートを選択してください:").pack()
        
        selected_sheet = tk.StringVar(value=sheet_names[0])
        for name in sheet_names:
            tk.Radiobutton(selection_window, text=name, variable=selected_sheet, value=name).pack(anchor=tk.W)
        
        def confirm_selection():
            if selected_sheet.get():
                selection_window.destroy()
            else:
                messagebox.showerror("エラー", "シートを選択してください")
        
        confirm_button = tk.Button(selection_window, text="OK", command=confirm_selection)
        confirm_button.pack()
        root.wait_window(selection_window)
        
        sheet = workbook[selected_sheet.get()]
    else:
        sheet = workbook[sheet_names[0]]
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    
    # 年・月とO37の値を取得
    year = sheet['C2'].value
    month = sheet['E2'].value
    cell_o37 = sheet['O37'].value
    return data, year, month, cell_o37

# 時間計算
def calculate_hours(time_str):
    try:
        start, end = map(int, time_str.split('-'))
        if end < start:
            end += 24  # 翌日にまたぐ場合
        hours = end - start
        # 休憩時間の調整
        if 6 <= hours < 7:
            hours -= 0.5
        elif 7 <= hours < 8:
            hours -= 0.75
        elif hours >= 8:
            hours -= 1
        return hours
    except:
        return 0

# v2.06追加出勤日数定義
def calculate_workdays(schedule_row):
    return sum(1 for day in schedule_row if isinstance(day, str) and re.match(r'\d{1,2}-\d{1,2}', day))  # XX-XX形式のセルをカウント

# 給与計算関数の修正
def calculate_salaries():
    try:
        base_wage = float(base_wage_entry.get())
        overtime_multiplier = float(overtime_multiplier_entry.get())
        adjustment_minutes = int(adjustment_entry.get())
        save_settings(base_wage, overtime_multiplier, adjustment_minutes)  # 設定を保存
        
        total_sum = 0
        total_hours_sum = 0
        total_workdays_sum = 0
        
        for i, name in enumerate(names):
            wage = float(wage_entries[i].get())
            total_hours = 0
            overtime_hours = 0
            workdays = calculate_workdays(schedule[i])
            adjustment_hours = (adjustment_minutes * workdays) / 60 if adjustment_checks[i].get() else 0
            
            for day in range(len(schedule[i])):
                hours = calculate_hours(schedule[i][day])
                total_hours += hours
                if hours > 8:
                    overtime_hours += (hours - 8)

            total_hours += adjustment_hours
            
            normal_pay = wage * (total_hours - overtime_hours)
            overtime_pay = wage * overtime_multiplier * overtime_hours
            total_salary = math.ceil(normal_pay + overtime_pay)
            
            salary_labels[i].config(text=f"{total_salary}円")
            hour_labels[i].config(text=f"{total_hours:.2f}時間")
            overtime_labels[i].config(text=f"{overtime_hours}時間")
            workdays_labels[i].config(text=f"{workdays}日")
            
            total_sum += total_salary
            total_hours_sum += total_hours
            total_workdays_sum += workdays

        total_label.config(text=f"{total_sum}円")
        total_hours_label.config(text=f"{total_hours_sum:.2f}時間")
    except ValueError:
        messagebox.showerror("エラー", "時給と倍率には数値を入力してください。")

# CSV保存の修正
def save_results():
    if year is None or month is None:
        messagebox.showerror("エラー", "年と月のデータが読み込まれていません")
        return
    
    file_name = f"{cell_o37}_{year}年{month}月16日～{month+1}月15日.csv".replace(".0", "")
    file_path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=file_name)
    
    if file_path:
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:  # UTF-8 with BOM
            writer = csv.writer(f)
            writer.writerow([cell_o37])
            writer.writerow([f"{year}年", f"{month}月16日", "～", f"{month+1}月15日"])
            writer.writerow(["基本時給：", f"{base_wage_entry.get()}円", "倍率：", f"{overtime_multiplier_entry.get()}倍"])
            writer.writerow([])
            writer.writerow(["名前", "基本時給", "出勤時間", "時間外", "出勤日数", "給与"])
            
            total_sum = 0
            total_hours_sum = 0
            total_workdays_sum = 0
            
            for i, name in enumerate(names):
                salary = salary_labels[i].cget("text")
                hours = hour_labels[i].cget("text")
                overtime_hours = overtime_labels[i].cget("text")
                workdays = workdays_labels[i].cget("text")
                
                writer.writerow([name, f"{wage_entries[i].get()}円", hours, overtime_hours, workdays, salary])
                
                total_sum += float(salary.replace("円", ""))
                total_hours_sum += float(hours.replace("時間", ""))
                total_workdays_sum += int(workdays.replace("日", ""))
            
            writer.writerow([])
            writer.writerow(["合計", "--", f"{total_hours_sum:.2f}時間", "--", "--",f"{total_sum:.0f}円"])

# ヘルプを開く
def open_help():
    webbrowser.open("https://sites.google.com/view/yo-payroll/")

#ソフトを終了する
def end_app():
    sys.exit()

# ソフトを再起動する
def restart_app():
    """アプリを再起動する"""
    root.destroy()  # Tkinterのウィンドウを閉じる
    python = sys.executable  # 現在のPython実行ファイル（exe化した場合も対応）
    os.execv(python, [python] + sys.argv)

# 設定を読み込み
settings = load_settings()
base_wage_default = settings.get("base_wage", 1162) if settings else 1162
overtime_multiplier_default = settings.get("overtime_multiplier", 1.25) if settings else 1.25

# ヘルプボタン
help_button = tk.Button(root, text="ヘルプ", command=open_help)
help_button.grid(row=0, column=6, sticky="e")

#閉じるボタン
end_button = tk.Button(root, text="✕", command=end_app)
end_button.grid(row=0, column=0, sticky="w")

# 上部の設定エリア
tk.Label(root, text="基本時給:").grid(row=1, column=0, sticky="e")
base_wage_entry = tk.Entry(root, width=10)
base_wage_entry.insert(0, base_wage_default)
base_wage_entry.grid(row=1, column=1, sticky="w")

tk.Label(root, text="時間外倍率:").grid(row=1, column=2, sticky="e")
overtime_multiplier_entry = tk.Entry(root, width=10)
overtime_multiplier_entry.insert(0, overtime_multiplier_default)
overtime_multiplier_entry.grid(row=1, column=3, sticky="w")

# 調整時間の設定（デフォルト20分）
tk.Label(root, text="調整(分):").grid(row=1, column=4, sticky="e")
adjustment_entry = tk.Entry(root, width=10)
adjustment_entry.insert(0, settings.get("adjustment_minutes", 20) if settings else 20)
adjustment_entry.grid(row=1, column=5, sticky="w")



# 調整時間ここまで

def apply_base_wage():
    wage = base_wage_entry.get()
    for entry in wage_entries:
        entry.delete(0, tk.END)
        entry.insert(0, wage)

apply_button = tk.Button(root, text="一括設定", command=apply_base_wage)
apply_button.grid(row=1, column=6, sticky="e")

name_text = tk.Label(root, text="　名前　", font=("",12,"bold"))
name_text.grid(row=2, column=0)

h_wage_text = tk.Label(root, text="基本時給", font=("",12,"bold"))
h_wage_text.grid(row=2, column=1)

add_text = tk.Label(root, text="　調整　", font=("",12,"bold"))
add_text.grid(row=2, column=2)

worktime_text = tk.Label(root, text="勤務時間", font=("",12,"bold"))
worktime_text.grid(row=2, column=3)

overtime_text = tk.Label(root, text=" 時間外 ", font=("",12,"bold"))
overtime_text.grid(row=2, column=4)

workdays_text = tk.Label(root, text="出勤日数", font=("",12,"bold"))
workdays_text.grid(row=2, column=5)

wage_text = tk.Label(root, text="　給与　", font=("",12,"bold"))
wage_text.grid(row=2, column=6)


# ファイル選択とデータ読み込み
file_path = filedialog.askopenfilename(title="シフト表ファイルを選択", filetypes=[("Excelファイル", "*.xlsx")], initialdir="C:/payroll_app/download")
if not file_path:
    sys.exit()

data, year, month, cell_o37 = load_sheet(file_path)
if data is None:
    sys.exit()

# ファイル情報の表示
file_info = f"{year}年{month}月16日～{month+1}月15日, {cell_o37}".replace(".0", "")
tk.Label(root, text=file_info, font=("",12,"bold")).grid(row=0, column=0, columnspan=7)

# 名前とシフトデータの抽出
names = []
schedule = []

header = data[2]
for col in range(3, len(header)):
    name = header[col]
    if name not in EXCLUDED_COLUMNS and name:
        names.append(name)
        schedule.append([row[col] for row in data[3:]])

# 従業員ごとの設定エリア
wage_entries = []
salary_labels = []
hour_labels = []
overtime_labels = []

for i, name in enumerate(names):
    tk.Label(root, text=name).grid(row=i+3, column=0)
    
    wage_entry = tk.Entry(root, width=10)
    wage_entry.insert(0, base_wage_entry.get())
    wage_entry.grid(row=i+3, column=1)
    wage_entries.append(wage_entry)
    
    hour_label = tk.Label(root, text="0時間")
    hour_label.grid(row=i+3, column=3)
    hour_labels.append(hour_label)
    
    overtime_label = tk.Label(root, text="0時間")
    overtime_label.grid(row=i+3, column=4)
    overtime_labels.append(overtime_label)
    
    salary_label = tk.Label(root, text="0円")
    salary_label.grid(row=i+3, column=6)
    salary_labels.append(salary_label)

# v2.06出勤日数追加
workdays_labels = []
for i, name in enumerate(names):
    workdays_label = tk.Label(root, text="0日")
    workdays_label.grid(row=i+3, column=5)
    workdays_labels.append(workdays_label)

    adjustment_checks = []
for i, name in enumerate(names):
    adjustment_var = tk.BooleanVar(value=True)
    adjustment_check = tk.Checkbutton(root, variable=adjustment_var)
    adjustment_check.grid(row=i+3, column=2)
    adjustment_checks.append(adjustment_var)

total_workdays_label = tk.Label(root, text="0日", font=("",12,"bold"))
total_workdays_label.grid(row=len(names)+3, column=2)

# 下部のボタンエリア
BG_label = tk.Label(root, text="", font=("",0,"bold"), bg= ("cyan1"))
BG_label.grid(row=len(names)+3, column=0, columnspan=7, sticky = tk.W+tk.E)

total_name_label = tk.Label(root, text="合計", font=("",12,"bold"), bg= ("cyan1"))
total_name_label.grid(row=len(names)+3, column=0)

total_warge_label = tk.Label(root, text="----", font=("",12,"bold"), bg= ("cyan1"))
total_warge_label.grid(row=len(names)+3, column=1, sticky = tk.W+tk.E)

total_hours_label = tk.Label(root, text="0時間", font=("",12,"bold"), bg= ("cyan1"))
total_hours_label.grid(row=len(names)+3, column=3)

total_over_label = tk.Label(root, text="----", font=("",12,"bold"), bg= ("cyan1"))
total_over_label.grid(row=len(names)+3, column=4)

total_workdays_label = tk.Label(root, text="----", font=("",12,"bold"), bg= ("cyan1"))
total_workdays_label.grid(row=len(names)+3, column=5)

total_add_label = tk.Label(root, text="----", font=("",12,"bold"), bg= ("cyan1"))
total_add_label.grid(row=len(names)+3, column=2)

total_label = tk.Label(root, text="0円", font=("",12,"bold"), bg= ("cyan1"))
total_label.grid(row=len(names)+3, column=6)

calculate_button = tk.Button(root, text="計算開始", command=calculate_salaries)
calculate_button.grid(row=len(names)+5, column=1)

save_button = tk.Button(root, text="ファイル保存", command=save_results)
save_button.grid(row=len(names)+5, column=3)

restart_button = tk.Button(root, text="はじめから", command=restart_app)
restart_button.grid(row=len(names)+5, column=5)
restart_button['state'] = 'disabled'

end_text = tk.Label(root, text="ボタンの上にカーソルを乗せると説明が表示されます。")
end_text.grid(row=len(names)+6, column=0, columnspan=7)

ver_text = tk.Label(root, text=f"ver.{APP_VERSION}")
ver_text.grid(row=len(names)+6, column=6, sticky="e")



#説明テキスト
Tooltip(help_button, "ヘルプページを開きます。")
Tooltip(base_wage_entry, "基本時給を入力してください。")
Tooltip(overtime_multiplier_entry, "時間外勤務の倍率を入力してください。")
Tooltip(adjustment_entry, "10分前出勤や交代時に生じるシフト表の時間以外の出勤時間を分単位で入力してください。")
Tooltip(apply_button, "全員の時給を一括で設定します。")
Tooltip(calculate_button, "計算を開始します。\n基本時給と時間外倍率、調整用の値が保存されます。")
Tooltip(save_button, "給与計算結果をCSVファイルとして保存します。")
Tooltip(name_text, "従業員名です。\n意図しない名前が表示されている場合は時給を0円に設定してください。")
Tooltip(h_wage_text, "基本時給です。\n早朝・深夜時給に対応していない為、個別で時給を入力してください。")
Tooltip(worktime_text, "1ヶ月あたりの勤務時間です。\n休憩時間は自動で引かれています。")
Tooltip(overtime_text, "1ヶ月あたりの時間外勤務時間です。\n8時間を超えた勤務の場合にカウントされます。")
Tooltip(workdays_text, "1ヶ月あたりの出勤日数です。\n8時間を超えた勤務の場合にカウントされます。")
Tooltip(add_text, "調整を適用するかのチェックボックスです。\n調整(分)に入力された時間×出勤日数が勤務時間に加算されます。\n給与手渡しの場合など、シフト表上以外の出勤時間が発生しない従業員のチェックは外して下さい。")
Tooltip(wage_text, "1ヶ月あたりの給与です。\n小数点以下は切り上げています。")
Tooltip(ver_text, "制作：熱田隆志\n最終更新：2025年03月03日")
Tooltip(end_button, "ソフトを終了します。")
Tooltip(restart_button, "この機能は現在メンテナンス中です。\n修正が完了するまでお待ち下さい。")

root.mainloop()